import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import requests
import time
import csv
from bs4 import BeautifulSoup
from datetime import datetime
import webbrowser
import threading
import os
import socket
import re
from urllib.parse import urlparse
import subprocess
import sys
import openpyxl
from openpyxl import Workbook

class WikiCheckApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WikiCheck - Проверка доменов в Wikipedia")
        self.root.geometry("700x500")
        
        # Заголовки для запросов
        self.headers = {
            "User-Agent": "Mozilla/5.0 (compatible; DropDomainBot/1.0; +mailto:veprik8900@mail.ru)"
        }
        
        self.output_path = ""
        self.setup_ui()
        
    def setup_ui(self):
        # Основной фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Заголовок
        title_label = ttk.Label(main_frame, text="WikiCheck - Поиск обратных ссылок с Wikipedia", 
                               font=("Arial", 14, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 10))
        
        # Кнопка выбора файла
        ttk.Label(main_frame, text="Файл с доменами:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.file_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.file_path, width=50).grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(main_frame, text="Выбрать", command=self.select_file).grid(row=1, column=2, pady=5)
        
        # Выбор пути сохранения
        ttk.Label(main_frame, text="Путь сохранения:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.output_path_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.output_path_var, width=50).grid(row=2, column=1, padx=5, pady=5)
        ttk.Button(main_frame, text="Выбрать", command=self.select_output_path).grid(row=2, column=2, pady=5)
        
        # Выбор формата
        ttk.Label(main_frame, text="Формат:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.format_var = tk.StringVar(value="Excel")
        format_combo = ttk.Combobox(main_frame, textvariable=self.format_var, values=["Excel", "CSV"], state="readonly")
        format_combo.grid(row=3, column=1, sticky=tk.W, padx=5, pady=5)
        
        # Кнопка запуска
        self.start_button = ttk.Button(main_frame, text="Начать проверку", command=self.start_check)
        self.start_button.grid(row=4, column=0, columnspan=3, pady=10)
        
        # Прогресс бар
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # Статус
        self.status_label = ttk.Label(main_frame, text="Готов к работе")
        self.status_label.grid(row=6, column=0, columnspan=3, pady=5)
        
        # Логи
        ttk.Label(main_frame, text="Логи:").grid(row=7, column=0, sticky=tk.W, pady=(10, 0))
        self.log_text = scrolledtext.ScrolledText(main_frame, width=80, height=12)
        self.log_text.grid(row=8, column=0, columnspan=3, pady=5, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Кнопка связи
        contact_button = ttk.Button(main_frame, text="Связаться", 
                                   command=self.open_contact)
        contact_button.grid(row=9, column=0, columnspan=3, pady=10)
        
        # Настройка растягивания
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(8, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Выберите файл с доменами",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        if file_path:
            self.file_path.set(file_path)
            
    def select_output_path(self):
        path = filedialog.askdirectory(title="Выберите папку для сохранения")
        if path:
            self.output_path_var.set(path)
            
    def open_contact(self):
        webbrowser.open("https://t.me/Userspoi")
        
    def log(self, message):
        self.log_text.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def validate_domain(self, domain):
        """
        Проверка валидности домена
        """
        domain_regex = re.compile(
            r"^[a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.([a-zA-Z]{2,}\.?)+$"
        )
        return domain_regex.match(domain) is not None
    
    def check_domain_exists(self, domain):
        """
        Проверка существования домена через DNS
        """
        try:
            socket.gethostbyname(domain)
            return True
        except socket.gaierror:
            return False
    
    def check_website_status(self, domain):
        """
        Проверка доступности сайта
        """
        try:
            response = requests.get(f"http://{domain}", timeout=5, allow_redirects=True)
            return response.status_code
        except:
            try:
                response = requests.get(f"https://{domain}", timeout=5, allow_redirects=True)
                return response.status_code
            except:
                return None
    
    def get_whois_info(self, domain):
        """
        Получение WHOIS информации через системную команду
        """
        try:
            if sys.platform == "win32":
                # Для Windows используем nslookup как альтернативу
                result = subprocess.run(['nslookup', domain], 
                                      capture_output=True, text=True, timeout=10,
                                      creationflags=subprocess.CREATE_NO_WINDOW)
                if result.returncode == 0 and result.stdout:
                    return {'status': 'Active', 'method': 'DNS lookup'}
            else:
                # Для Unix/Linux используем системную команду whois
                result = subprocess.run(['whois', domain], 
                                      capture_output=True, text=True, timeout=10)
                if result.returncode == 0 and result.stdout:
                    # Простой парсинг основных данных
                    output = result.stdout.lower()
                    info = {'status': 'Active', 'method': 'System whois'}
                    
                    # Попытка извлечь основную информацию
                    if 'creation date' in output or 'created' in output:
                        info['has_creation_date'] = True
                    if 'registrar' in output:
                        info['has_registrar'] = True
                    
                    return info
            
            return None
        except Exception as e:
            self.log(f"    ⚠️ WHOIS недоступен для {domain}: {e}")
            return None
    
    def analyze_domain_flags(self, domain_info, website_status):
        """
        Анализ красных флагов домена
        """
        flags = []
        
        # Проверка доступности
        if website_status is None:
            flags.append("Сайт недоступен")
        elif website_status >= 400:
            flags.append(f"HTTP ошибка {website_status}")
        
        # Проверка WHOIS данных
        if not domain_info:
            flags.append("WHOIS недоступен")
        
        return flags
    
    def search_wikipedia_links(self, domain):
        """
        Использует Bing для поиска ссылок с Wikipedia на указанный домен
        """
        try:
            query = f"site:en.wikipedia.org OR site:ru.wikipedia.org {domain}"
            url = f"https://www.bing.com/search?q={query}"
            response = requests.get(url, headers=self.headers, timeout=10)

            soup = BeautifulSoup(response.text, 'html.parser')
            results = []

            for li in soup.find_all("li", class_="b_algo"):
                link = li.find("a")
                if link and "wikipedia.org" in link['href']:
                    title = link.get_text(strip=True)
                    href = link['href']
                    results.append((href, title))
            return results
        except Exception as e:
            self.log(f"    ❌ Ошибка при поиске для {domain}: {e}")
            return []
    
    def process_domain(self, domain):
        """
        Полная обработка домена с проверками
        """
        # Базовая информация о домене
        domain_data = {
            'domain': domain,
            'date': datetime.now().strftime('%Y-%m-%d'),
            'valid_format': False,
            'dns_exists': False,
            'website_status': None,
            'whois_available': False,
            'wikipedia_links': [],
            'flags': []
        }
        
        # 1. Валидация домена
        if not self.validate_domain(domain):
            self.log(f"    ❌ Некорректный формат домена: {domain}")
            domain_data['flags'].append("Некорректный формат")
            return domain_data
        
        domain_data['valid_format'] = True
        
        # 2. Проверка существования
        if not self.check_domain_exists(domain):
            self.log(f"    ❌ Домен не найден в DNS: {domain}")
            domain_data['flags'].append("Не найден в DNS")
            return domain_data
        
        domain_data['dns_exists'] = True
        self.log(f"    ✅ Домен найден в DNS")
        
        # 3. Проверка доступности сайта
        website_status = self.check_website_status(domain)
        domain_data['website_status'] = website_status
        
        if website_status:
            self.log(f"    ✅ Сайт доступен (HTTP {website_status})")
        else:
            self.log(f"    ❌ Сайт недоступен")
        
        # 4. Получение WHOIS информации
        self.log(f"    🔍 Получение WHOIS данных...")
        domain_info = self.get_whois_info(domain)
        
        if domain_info:
            domain_data['whois_available'] = True
            self.log(f"    ✅ WHOIS данные получены")
            if domain_info.get('has_registrar'):
                self.log(f"    🏢 Регистратор найден")
        else:
            self.log(f"    ⚠️ WHOIS данные недоступны")
        
        # 5. Анализ красных флагов
        flags = self.analyze_domain_flags(domain_info, website_status)
        domain_data['flags'].extend(flags)
        
        if flags:
            self.log(f"    🚩 Красные флаги: {', '.join(flags)}")
        
        # 6. Поиск ссылок в Wikipedia
        self.log(f"    🔍 Поиск ссылок в Wikipedia...")
        links = self.search_wikipedia_links(domain)
        domain_data['wikipedia_links'] = links
        
        if links:
            for url, anchor in links:
                self.log(f"    ✅ Найдена ссылка: {url}")
        else:
            self.log(f"    ❌ Ссылок не найдено")
        
        return domain_data
    
    def save_results(self, all_results):
        """
        Сохранение результатов в выбранном формате
        """
        if not self.output_path_var.get():
            output_dir = os.path.dirname(os.path.abspath(__file__))
        else:
            output_dir = self.output_path_var.get()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if self.format_var.get() == "Excel":
            filename = f"wikipedia_check_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            self.save_to_excel(all_results, filepath)
        else:
            filename = f"wikipedia_check_{timestamp}.csv"
            filepath = os.path.join(output_dir, filename)
            self.save_to_csv(all_results, filepath)
        
        return filepath
    
    def save_to_excel(self, all_results, filepath):
        """
        Сохранение в Excel
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Wikipedia Check Results"
        
        # Заголовки
        headers = [
            'Домен', 'Дата проверки', 'Валидный формат', 'DNS существует', 
            'HTTP статус', 'WHOIS доступен', 'Количество Wiki-ссылок', 
            'Wiki-ссылки', 'Тексты ссылок', 'Красные флаги'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Данные
        row = 2
        for domain_data in all_results:
            ws.cell(row=row, column=1, value=domain_data['domain'])
            ws.cell(row=row, column=2, value=domain_data['date'])
            ws.cell(row=row, column=3, value='Да' if domain_data['valid_format'] else 'Нет')
            ws.cell(row=row, column=4, value='Да' if domain_data['dns_exists'] else 'Нет')
            ws.cell(row=row, column=5, value=domain_data['website_status'] if domain_data['website_status'] else 'N/A')
            ws.cell(row=row, column=6, value='Да' if domain_data['whois_available'] else 'Нет')
            ws.cell(row=row, column=7, value=len(domain_data['wikipedia_links']))
            
            # Ссылки и тексты
            if domain_data['wikipedia_links']:
                urls = [link[0] for link in domain_data['wikipedia_links']]
                texts = [link[1] for link in domain_data['wikipedia_links']]
                ws.cell(row=row, column=8, value='\n'.join(urls))
                ws.cell(row=row, column=9, value='\n'.join(texts))
            else:
                ws.cell(row=row, column=8, value='Нет')
                ws.cell(row=row, column=9, value='Нет')
            
            ws.cell(row=row, column=10, value='; '.join(domain_data['flags']) if domain_data['flags'] else 'Нет')
            row += 1
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def save_to_csv(self, all_results, filepath):
        """
        Сохранение в CSV
        """
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Домен', 'Дата проверки', 'Валидный формат', 'DNS существует', 
                'HTTP статус', 'WHOIS доступен', 'Количество Wiki-ссылок', 
                'Wiki-ссылки', 'Тексты ссылок', 'Красные флаги'
            ])
            
            for domain_data in all_results:
                urls = '|'.join([link[0] for link in domain_data['wikipedia_links']]) if domain_data['wikipedia_links'] else 'Нет'
                texts = '|'.join([link[1] for link in domain_data['wikipedia_links']]) if domain_data['wikipedia_links'] else 'Нет'
                
                writer.writerow([
                    domain_data['domain'],
                    domain_data['date'],
                    'Да' if domain_data['valid_format'] else 'Нет',
                    'Да' if domain_data['dns_exists'] else 'Нет',
                    domain_data['website_status'] if domain_data['website_status'] else 'N/A',
                    'Да' if domain_data['whois_available'] else 'Нет',
                    len(domain_data['wikipedia_links']),
                    urls,
                    texts,
                    '; '.join(domain_data['flags']) if domain_data['flags'] else 'Нет'
                ])
    
    def start_check(self):
        if not self.file_path.get():
            messagebox.showerror("Ошибка", "Выберите файл с доменами")
            return
            
        if not os.path.exists(self.file_path.get()):
            messagebox.showerror("Ошибка", "Файл не найден")
            return
            
        # Запускаем проверку в отдельном потоке
        self.start_button.config(state='disabled')
        thread = threading.Thread(target=self.check_domains)
        thread.daemon = True
        thread.start()
        
    def check_domains(self):
        try:
            # Читаем домены из файла
            with open(self.file_path.get(), 'r', encoding='utf-8') as f:
                domains = [line.strip() for line in f if line.strip()]
            
            if not domains:
                messagebox.showerror("Ошибка", "Файл пуст или не содержит доменов")
                self.start_button.config(state='normal')
                return
            
            self.log(f"Загружено {len(domains)} доменов для проверки")
            
            # Настраиваем прогресс бар
            self.progress.config(maximum=len(domains))
            self.progress.config(value=0)
            
            all_results = []
            wiki_links_found = 0
            
            for i, domain in enumerate(domains, 1):
                self.status_label.config(text=f"Проверяю: {domain} ({i}/{len(domains)})")
                self.log(f"[{i}/{len(domains)}] Проверяю: {domain}")
                
                # Обработка домена
                domain_data = self.process_domain(domain)
                all_results.append(domain_data)
                
                if domain_data['wikipedia_links']:
                    wiki_links_found += len(domain_data['wikipedia_links'])
                
                # Обновляем прогресс
                self.progress.config(value=i)
                self.root.update()
                
                time.sleep(2)  # пауза между запросами
            
            # Сохраняем все результаты
            output_file = self.save_results(all_results)
            
            self.log(f"\n✅ Проверено доменов: {len(all_results)}")
            self.log(f"✅ Найдено ссылок с Wikipedia: {wiki_links_found}")
            self.log(f"Результаты сохранены в {output_file}")
            self.status_label.config(text=f"Готово! Проверено {len(all_results)} доменов")
            messagebox.showinfo("Готово", f"Проверено {len(all_results)} доменов.\nНайдено {wiki_links_found} ссылок с Wikipedia.\nРезультаты сохранены в {os.path.basename(output_file)}")
                
        except Exception as e:
            self.log(f"❌ Общая ошибка: {e}")
            messagebox.showerror("Ошибка", f"Произошла ошибка: {e}")
        finally:
            self.start_button.config(state='normal')
            self.progress.config(value=0)

def main():
    root = tk.Tk()
    app = WikiCheckApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
